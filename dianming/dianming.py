import random
import os
import json
import csv
from datetime import datetime
import tkinter as tk
from tkinter import messagebox, scrolledtext, filedialog

# 尝试导入 openpyxl，用于 Excel 读取与导出
try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# 默认配置（英文键，首次运行时自动创建）
DEFAULT_CONFIG = {
    "window_title": "随机点名程序",
    "window_size": [500, 700],
    "resizable": True,
    "animation_steps": 15,
    "animation_delay": 50,
    "default_weight": 1.0,
    "students": [
        {"name": "张三", "weight": 1.0},
        {"name": "李四", "weight": 1.0},
        {"name": "王五", "weight": 1.0},
        {"name": "张一", "weight": 1.0},
        {"name": "张二", "weight": 1.0}
    ]
}

class RandomNameApp:
    def __init__(self, root, config):
        self.root = root
        self.config = config

        # 应用窗口配置
        self.root.title(config["window_title"])
        w, h = config["window_size"]
        self.root.geometry(f"{w}x{h}")
        self.root.resizable(config["resizable"], config["resizable"])

        self.all_names = []       # 全部姓名
        self.remaining = []       # 未被抽过的名单（不重复抽取）
        self.weights = {}         # 每位学生的权重
        self.history = []         # 点名历史记录：每项 {'time': str, 'name': str, 'mode': str}

        self.create_widgets()
        self.load_students()

    # ----------------- 界面布局 -----------------
    def create_widgets(self):
        # 顶部框架：标题（左） + 关于按钮（右）
        top_frame = tk.Frame(self.root)
        top_frame.pack(fill=tk.X, pady=(15, 0))

        title = tk.Label(top_frame, text="随机点名", font=("微软雅黑", 24, "bold"), fg="#2c3e50")
        title.pack(side=tk.TOP, padx=10)

        about_btn = tk.Button(top_frame, text="关于", font=("微软雅黑", 9),
                              command=self.show_about, bg="#95a5a6", fg="white", relief=tk.FLAT)
        about_btn.pack(side=tk.RIGHT, padx=10)

        # 结果显示
        self.result_var = tk.StringVar()
        self.result_var.set("就绪")
        result_label = tk.Label(self.root, textvariable=self.result_var,
                                font=("微软雅黑", 56, "bold"), fg="#e74c3c",
                                bg="#ecf0f1", width=12, height=2, relief="ridge")
        result_label.pack(pady=10)

        # 统计信息
        info_frame = tk.Frame(self.root)
        info_frame.pack(pady=5)
        self.total_var = tk.StringVar()
        self.left_var = tk.StringVar()
        tk.Label(info_frame, text="总人数：", font=("微软雅黑", 12)).grid(row=0, column=0, padx=5)
        tk.Label(info_frame, textvariable=self.total_var, font=("微软雅黑", 12, "bold")).grid(row=0, column=1, padx=5)
        tk.Label(info_frame, text="剩余未抽：", font=("微软雅黑", 12)).grid(row=0, column=2, padx=15)
        tk.Label(info_frame, textvariable=self.left_var, font=("微软雅黑", 12, "bold")).grid(row=0, column=3, padx=5)

        # 按钮区域
        btn_frame = tk.Frame(self.root)
        btn_frame.pack(pady=15)

        tk.Button(btn_frame, text="随机点名（不重复）", width=20, height=2,
                  command=self.pick_random, bg="#3498db", fg="white",
                  font=("微软雅黑", 11)).grid(row=0, column=0, padx=10, pady=5)

        tk.Button(btn_frame, text="独立随机（可重复）", width=20, height=2,
                  command=self.pick_random_independent, bg="#2ecc71", fg="white",
                  font=("微软雅黑", 11)).grid(row=0, column=1, padx=10, pady=5)

        tk.Button(btn_frame, text="查看剩余名单", width=20, height=2,
                  command=self.show_remaining, bg="#f39c12", fg="white",
                  font=("微软雅黑", 11)).grid(row=1, column=0, padx=10, pady=5)

        tk.Button(btn_frame, text="重置名单", width=20, height=2,
                  command=self.reset_pool, bg="#9b59b6", fg="white",
                  font=("微软雅黑", 11)).grid(row=1, column=1, padx=10, pady=5)

        tk.Button(btn_frame, text="从 Excel 导入", width=20, height=2,
                  command=self.import_from_excel, bg="#1abc9c", fg="white",
                  font=("微软雅黑", 11)).grid(row=2, column=0, columnspan=2, pady=5)

        # 查看历史与导出历史放在同一行
        tk.Button(btn_frame, text="查看历史记录", width=20, height=2,
                  command=self.show_history, bg="#f1c40f", fg="white",
                  font=("微软雅黑", 11)).grid(row=3, column=0, padx=10, pady=5)

        tk.Button(btn_frame, text="导出历史记录", width=20, height=2,
                  command=self.export_history, bg="#e67e22", fg="white",
                  font=("微软雅黑", 11)).grid(row=3, column=1, padx=10, pady=5)

        tk.Button(self.root, text="退出程序", width=15, height=2,
                  command=self.root.quit, bg="#e74c3c", fg="white",
                  font=("微软雅黑", 11)).pack(pady=10)

    # ----------------- 关于对话框 -----------------
    def show_about(self):
        messagebox.showinfo("关于随机点名程序",
                            "随机点名程序 v1.2\n\n"
                            "一个支持权重调节、Excel导入的随机点名工具\n"
                            "可记录、查看并导出点名历史\n\n"
                            "配置文件：config.json\n"
                            "名单与权重可直接编辑该文件，也可通过Excel导入\n\n"
                            "AI制作, 作者: github.com/temingyu\n\n"
                            "(作者没啥实力只会用AI了( )")

    # ----------------- 学生数据加载 -----------------
    def load_students(self):
        students = self.config.get("students", [])
        if not students:
            messagebox.showerror("配置错误", "配置文件中没有学生数据，请检查 config.json 中的 students 字段。")
            self.root.quit()
            return

        self.all_names = []
        self.weights = {}
        for stu in students:
            name = stu.get("name", "").strip()
            weight = stu.get("weight", 1.0)
            if not name:
                continue
            if not isinstance(weight, (int, float)) or weight < 0:
                weight = 1.0
            self.all_names.append(name)
            self.weights[name] = float(weight)

        if not self.all_names:
            messagebox.showerror("配置错误", "名单中没有有效姓名。")
            self.root.quit()
            return

        self.remaining = self.all_names.copy()
        self.update_status()

    # ----------------- Excel 导入 -----------------
    def import_from_excel(self):
        if not HAS_OPENPYXL:
            messagebox.showerror("缺少库", "需要安装 openpyxl 库才能读取 Excel 文件。\n\n请执行：pip install openpyxl")
            return

        file_path = filedialog.askopenfilename(
            title="选择 Excel 文件",
            filetypes=[("Excel 文件", "*.xlsx"), ("所有文件", "*.*")]
        )
        if not file_path:
            return

        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
        except Exception as e:
            messagebox.showerror("读取失败", f"无法打开 Excel 文件：{e}")
            return

        new_students = []
        header_offset = 0
        if ws.max_row >= 1:
            first_cell_b = ws.cell(row=1, column=2).value
            if isinstance(first_cell_b, str) and not first_cell_b.replace('.', '', 1).isdigit():
                header_offset = 1

        for row in ws.iter_rows(min_row=1 + header_offset, values_only=True):
            if not row or len(row) < 1:
                continue
            name = str(row[0]).strip() if row[0] else ""
            if not name:
                continue
            weight = self.config.get("default_weight", 1.0)
            if len(row) >= 2 and row[1] is not None:
                try:
                    w = float(row[1])
                    if w >= 0:
                        weight = w
                except (ValueError, TypeError):
                    pass
            new_students.append({"name": name, "weight": weight})

        if not new_students:
            messagebox.showwarning("无数据", "在 Excel 文件中未找到任何有效姓名。")
            return

        save_to_config = messagebox.askyesno("导入成功",
                                             f"成功导入 {len(new_students)} 名学生。\n是否将新名单保存到 config.json？")
        if save_to_config:
            self.config["students"] = new_students
            try:
                with open("config.json", "w", encoding="utf-8") as f:
                    json.dump(self.config, f, ensure_ascii=False, indent=4)
            except Exception as e:
                messagebox.showerror("保存失败", f"无法写入 config.json：{e}")

        self.all_names = [s["name"] for s in new_students]
        self.weights = {s["name"]: s["weight"] for s in new_students}
        self.remaining = self.all_names.copy()
        self.update_status()
        self.result_var.set("就绪")

    # ----------------- 核心功能 -----------------
    def weighted_choice(self, name_list):
        if not name_list:
            return None
        w = [self.weights.get(name, 0.0) for name in name_list]
        if sum(w) == 0:
            return random.choice(name_list)
        return random.choices(name_list, weights=w, k=1)[0]

    def pick_random(self):
        if not self.remaining:
            messagebox.showinfo("提示", "所有人都已被抽过，请重置名单。")
            return
        self._animate_choice(from_pool=self.remaining, final_callback=self._remove_from_remaining, mode="不重复")

    def _remove_from_remaining(self, name):
        if name in self.remaining:
            self.remaining.remove(name)
        self.update_status()

    def pick_random_independent(self):
        if not self.all_names:
            messagebox.showinfo("提示", "名单为空，无法抽取。")
            return
        self._animate_choice(from_pool=self.all_names, final_callback=None, mode="独立随机")

    def reset_pool(self):
        self.remaining = self.all_names.copy()
        self.update_status()
        self.result_var.set("已重置")
        messagebox.showinfo("重置", "名单已重置，所有人均可再次抽取。")

    def show_remaining(self):
        if not self.remaining:
            messagebox.showinfo("剩余名单", "剩余名单为空。")
            return
        win = tk.Toplevel(self.root)
        win.title("剩余未抽名单")
        win.geometry("250x300")
        win.resizable(False, False)
        tk.Label(win, text="剩余未抽名单", font=("微软雅黑", 14, "bold")).pack(pady=10)
        text_area = scrolledtext.ScrolledText(win, width=25, height=12, font=("微软雅黑", 11))
        text_area.pack(padx=10, pady=5)
        for i, name in enumerate(self.remaining, 1):
            text_area.insert(tk.END, f"{i}. {name}\n")
        text_area.config(state=tk.DISABLED)

    def update_status(self):
        self.total_var.set(str(len(self.all_names)))
        self.left_var.set(str(len(self.remaining)))

    # ----------------- 历史记录查看 -----------------
    def show_history(self):
        """弹出窗口显示点名历史记录"""
        if not self.history:
            messagebox.showinfo("历史记录", "暂无点名历史记录。")
            return
        win = tk.Toplevel(self.root)
        win.title("点名历史记录")
        win.geometry("550x400")
        win.resizable(True, True)

        tk.Label(win, text="点名历史记录", font=("微软雅黑", 14, "bold")).pack(pady=10)

        # 使用 ScrolledText 显示表格样式的记录
        text_area = scrolledtext.ScrolledText(win, width=65, height=18, font=("微软雅黑", 10))
        text_area.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        # 插入表头
        text_area.insert(tk.END, f"{'序号':<6}{'时间':<22}{'姓名':<12}{'模式'}\n")
        text_area.insert(tk.END, "-" * 60 + "\n")

        for idx, record in enumerate(self.history, 1):
            line = f"{idx:<6}{record['time']:<22}{record['name']:<12}{record['mode']}\n"
            text_area.insert(tk.END, line)

        text_area.config(state=tk.DISABLED)

    # ----------------- 历史记录导出 -----------------
    def add_history(self, name, mode):
        """添加一条点名记录"""
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.history.append({"time": now, "name": name, "mode": mode})

    def export_history(self):
        """导出历史记录到 Excel 或 CSV 文件"""
        if not self.history:
            messagebox.showinfo("无记录", "暂无点名历史记录。")
            return

        if HAS_OPENPYXL:
            file_path = filedialog.asksaveasfilename(
                title="导出历史记录",
                defaultextension=".xlsx",
                filetypes=[("Excel 文件", "*.xlsx"), ("CSV 文件", "*.csv"), ("所有文件", "*.*")]
            )
        else:
            file_path = filedialog.asksaveasfilename(
                title="导出历史记录",
                defaultextension=".csv",
                filetypes=[("CSV 文件", "*.csv"), ("所有文件", "*.*")]
            )

        if not file_path:
            return

        try:
            if file_path.endswith('.xlsx') and HAS_OPENPYXL:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "点名历史"
                ws.append(["序号", "时间", "姓名", "模式"])
                for idx, record in enumerate(self.history, 1):
                    ws.append([idx, record["time"], record["name"], record["mode"]])
                wb.save(file_path)
            else:
                with open(file_path, 'w', newline='', encoding='utf-8-sig') as f:
                    writer = csv.writer(f)
                    writer.writerow(["序号", "时间", "姓名", "模式"])
                    for idx, record in enumerate(self.history, 1):
                        writer.writerow([idx, record["time"], record["name"], record["mode"]])

            messagebox.showinfo("导出成功", f"历史记录已保存到：\n{file_path}")
        except Exception as e:
            messagebox.showerror("导出失败", f"保存文件时发生错误：{e}")

    # ----------------- 动画效果 -----------------
    def _animate_choice(self, from_pool, final_callback, mode):
        steps = self.config["animation_steps"]
        delay = self.config["animation_delay"]
        if not from_pool:
            return
        final_choice = self.weighted_choice(from_pool)
        pool_copy = from_pool.copy() if len(from_pool) > 1 else from_pool
        self._animate_step(pool_copy, final_choice, final_callback, steps, delay, 0, mode)

    def _animate_step(self, pool, final_choice, final_callback, steps, delay, count, mode):
        if count < steps:
            temp = random.choice(pool)
            self.result_var.set(temp)
            self.root.after(delay, lambda: self._animate_step(pool, final_choice, final_callback, steps, delay, count + 1, mode))
        else:
            self.result_var.set(final_choice)
            self.add_history(final_choice, mode)
            if final_callback:
                final_callback(final_choice)

# ----------------- 配置加载（自动创建） -----------------
def load_config():
    config_path = "config.json"
    if not os.path.exists(config_path):
        try:
            with open(config_path, "w", encoding="utf-8") as f:
                json.dump(DEFAULT_CONFIG, f, ensure_ascii=False, indent=4)
            messagebox.showinfo("配置文件已创建",
                                f"已在当前目录下创建默认配置文件 {config_path}，\n您可以根据需要修改其中的 students 字段。")
        except Exception as e:
            messagebox.showerror("错误", f"无法创建配置文件 {config_path}：{e}")
            return DEFAULT_CONFIG.copy()

    try:
        with open(config_path, "r", encoding="utf-8") as f:
            config = json.load(f)
    except Exception as e:
        messagebox.showerror("配置文件错误", f"读取 {config_path} 失败：{e}\n将使用默认配置运行。")
        return DEFAULT_CONFIG.copy()

    # 用默认值补全可能缺失的键
    for key, value in DEFAULT_CONFIG.items():
        if key not in config:
            config[key] = value
    return config

# ----------------- 程序入口 -----------------
if __name__ == "__main__":
    config = load_config()
    root = tk.Tk()
    app = RandomNameApp(root, config)
    root.mainloop()